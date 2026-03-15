"""
Enhanced GST Details Import View with PAN-Based Validation
Follows the same workflow and structure as the CA Assignment module
Supports multiple GST numbers per client - only adds new GST numbers, never updates
"""
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.contrib import messages
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from io import BytesIO
import re

from home.models import Client, GSTDetails, STATE_CHOICES


def is_superuser_or_admin(user):
    """Check if user is superuser or admin"""
    if user.is_superuser:
        return True
    try:
        return hasattr(user, 'employee') and user.employee.role == 'ADMIN'
    except:
        return False


@login_required
@user_passes_test(is_superuser_or_admin, login_url='/accounts/login/')
def import_gst_view(request):
    """Main view for GST import page"""
    context = {
        'page_title': 'Import GST Details',
    }
    return render(request, 'gst_import/import_gst.html', context)


@login_required
@user_passes_test(is_superuser_or_admin, login_url='/accounts/login/')
def search_clients_for_gst_ajax(request):
    """AJAX endpoint to search clients by name or PAN for GST import"""
    q = request.GET.get('q', '').strip()
    
    if len(q) < 2:
        return JsonResponse([], safe=False)
    
    clients = Client.objects.filter(
        Q(client_name__icontains=q) | Q(pan_no__icontains=q)
    )[:20]
    
    data = []
    for c in clients:
        # Show existing GST count
        gst_count = c.gst_details.count()
        gst_info = f" ({gst_count} GST)" if gst_count > 0 else " (No GST)"
        
        data.append({
            "id": c.id,
            "text": f"{c.client_name} - {c.pan_no or 'No PAN'}{gst_info}"
        })
    
    return JsonResponse(data, safe=False)


@login_required
@user_passes_test(is_superuser_or_admin)
def bulk_import_gst(request):
    """Handle bulk GST import via Excel/CSV upload with enhanced PAN validation"""
    print(f"bulk_import_gst called with method: {request.method}")
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No file uploaded'})
    
    file = request.FILES['file']
    print(f"File uploaded: {file.name}, size: {file.size}")
    
    # Validate file extension
    if not (file.name.endswith('.xlsx') or file.name.endswith('.xls') or file.name.endswith('.csv')):
        return JsonResponse({
            'success': False,
            'message': 'Invalid file format. Please upload Excel or CSV file'
        })
    
    try:
        print("Starting file processing...")
        
        # Read file
        if file.name.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        print(f"File read successfully. Shape: {df.shape}")
        print(f"Columns: {list(df.columns)}")
        
        # Flexible column mapping (support both old and new formats)
        column_mapping = {}
        
        for col in df.columns:
            col_upper = col.upper().strip()
            if col_upper in ['CLIENT_NAME', 'CLIENT NAME', 'PROPRIETOR NAME']:
                column_mapping['client_name'] = col
            elif col_upper in ['PAN_NO', 'PAN NO', 'PAN NUMBER', 'PAN']:
                column_mapping['pan_no'] = col
            elif col_upper in ['GST_NUMBER', 'GST NO', 'GST NUMBER', 'GSTIN', 'GST_NO']:
                column_mapping['gst_number'] = col
            elif col_upper in ['REGISTERED_ADDRESS', 'ADDRESS']:
                column_mapping['registered_address'] = col
            elif col_upper in ['STATE_CODE', 'STATE CODE', 'STATE', 'STATE_NAME', 'STATE NAME']:
                column_mapping['state'] = col
            elif col_upper in ['GST_SCHEME_TYPE', 'SCHEME TYPE', 'SCHEME']:
                column_mapping['gst_scheme_type'] = col
            elif col_upper in ['STATUS']:
                column_mapping['status'] = col
        
        print(f"Column mapping: {column_mapping}")
        
        # Validate required columns
        required_columns = ['client_name', 'gst_number']
        missing_columns = [col for col in required_columns if col not in column_mapping]
        
        if missing_columns:
            missing_display = []
            for col in missing_columns:
                if col == 'client_name':
                    missing_display.append('Client Name')
                elif col == 'gst_number':
                    missing_display.append('GST Number')
            
            return JsonResponse({
                'success': False,
                'message': f'Missing required columns: {", ".join(missing_display)}'
            })
        
        # Process rows
        total_rows = len(df)
        added_records = []
        skipped_records = []
        
        print(f"Processing {total_rows} rows...")
        
        # GST validation regex (15-character format)
        gst_pattern = re.compile(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$')
        
        # Get valid state codes from the model's STATE_CHOICES
        VALID_STATE_CODES = dict(STATE_CHOICES)
        print(f"Valid state codes from model: {list(VALID_STATE_CODES.keys())}")
        
        # GST State Code to State Name Mapping (for display purposes)
        GST_STATE_NAME_MAPPING = VALID_STATE_CODES
        
        # Valid scheme types and statuses
        VALID_SCHEMES = ['Regular', 'Composition', 'QRMP']
        VALID_STATUSES = ['Active', 'Closed']
        
        with transaction.atomic():
            for index, row in df.iterrows():
                print(f"Processing row {index + 1}...")
                
                # Extract data using column mapping
                client_name = str(row[column_mapping['client_name']]).strip() if pd.notna(row[column_mapping['client_name']]) else ''
                pan_no = str(row[column_mapping.get('pan_no', '')]).strip().upper() if column_mapping.get('pan_no') and pd.notna(row[column_mapping.get('pan_no', '')]) else ''
                gst_number = str(row[column_mapping['gst_number']]).strip().upper() if pd.notna(row[column_mapping['gst_number']]) else ''
                
                # Optional fields - Define state field here BEFORE using it in print
                registered_address = str(row[column_mapping.get('registered_address', '')]).strip() if column_mapping.get('registered_address') and pd.notna(row[column_mapping.get('registered_address', '')]) else ''
                state_input = str(row[column_mapping.get('state', '')]).strip() if column_mapping.get('state') and pd.notna(row[column_mapping.get('state', '')]) else ''
                gst_scheme_type = str(row[column_mapping.get('gst_scheme_type', '')]).strip() if column_mapping.get('gst_scheme_type') and pd.notna(row[column_mapping.get('gst_scheme_type', '')]) else 'Regular'
                status = str(row[column_mapping.get('status', '')]).strip() if column_mapping.get('status') and pd.notna(row[column_mapping.get('status', '')]) else 'Active'
                
                print(f"Row {index + 1}: client_name='{client_name}', pan_no='{pan_no}', gst_number='{gst_number}'")
                print(f"Row {index + 1}: GST state code (first 2 digits): {gst_number[:2] if len(gst_number) >= 2 else 'Invalid'}")
                print(f"Row {index + 1}: Provided state input: '{state_input}'")
                
                # Skip if required fields are missing
                if not client_name or not gst_number:
                    skipped_records.append({
                        'client_name': client_name,
                        'pan_no': pan_no,
                        'gst_number': gst_number,
                        'reason': 'Required fields missing (Client Name or GST Number)'
                    })
                    continue
                
                # Validate GST number format
                if len(gst_number) != 15 or not gst_pattern.match(gst_number):
                    skipped_records.append({
                        'client_name': client_name,
                        'pan_no': pan_no,
                        'gst_number': gst_number,
                        'reason': 'Invalid GST format (must be 15 characters)'
                    })
                    continue
                
                # Auto-derive state code from GST number (first 2 digits)
                state_code_from_gst = gst_number[:2]
                
                # Create reverse mapping for state name to state code lookup
                STATE_NAME_TO_CODE = {name: code for code, name in VALID_STATE_CODES.items()}
                
                # State validation and selection logic
                state_code = None  # Initialize state_code variable
                
                if not state_input:
                    # If state is blank, use state code from GST number
                    state_code = state_code_from_gst
                    print(f"State input blank, using from GST: {state_code}")
                else:
                    # Check if state_input is a state code (2 digits) or state name
                    if len(state_input) == 2 and state_input.isdigit():
                        # It's a state code
                        if state_input != state_code_from_gst:
                            skipped_records.append({
                                'client_name': client_name,
                                'pan_no': pan_no,
                                'gst_number': gst_number,
                                'reason': f'State code mismatch: GST has {state_code_from_gst} but provided {state_input}'
                            })
                            continue
                        state_code = state_input
                        print(f"State code provided and validated: {state_code}")
                    else:
                        # It's a state name, convert to state code
                        state_code = STATE_NAME_TO_CODE.get(state_input)
                        if not state_code:
                            skipped_records.append({
                                'client_name': client_name,
                                'pan_no': pan_no,
                                'gst_number': gst_number,
                                'reason': f'Invalid state name: {state_input} (not found in valid states)'
                            })
                            continue
                        
                        # Validate that the state name matches GST-derived state code
                        if state_code != state_code_from_gst:
                            skipped_records.append({
                                'client_name': client_name,
                                'pan_no': pan_no,
                                'gst_number': gst_number,
                                'reason': f'State name mismatch: GST has state code {state_code_from_gst} ({VALID_STATE_CODES[state_code_from_gst]}) but provided {state_input}'
                            })
                            continue
                        print(f"State name '{state_input}' converted to code {state_code} and validated against GST")
                
                # Validate state code exists in our STATE_CHOICES
                if state_code not in VALID_STATE_CODES:
                    skipped_records.append({
                        'client_name': client_name,
                        'pan_no': pan_no,
                        'gst_number': gst_number,
                        'reason': f'Invalid state code: {state_code} (not in valid STATE_CHOICES: {list(VALID_STATE_CODES.keys())})'
                    })
                    continue
                
                print(f"State code {state_code} is valid, state name: {VALID_STATE_CODES[state_code]}")
                
                # Get the state name for display purposes
                state_name = GST_STATE_NAME_MAPPING.get(state_code, f'State Code {state_code}')
                
                # Validate GST scheme type
                if gst_scheme_type not in VALID_SCHEMES:
                    gst_scheme_type = 'Regular'  # Default to Regular
                
                # Validate status
                if status not in VALID_STATUSES:
                    status = 'Active'  # Default to Active
                
                # Enhanced Client Matching Logic with PAN validation
                client = None
                
                if pan_no:
                    # Primary matching: Both Client Name and PAN No (most accurate)
                    try:
                        client = Client.objects.get(
                            client_name__iexact=client_name,
                            pan_no__iexact=pan_no
                        )
                        print(f"Found client by name+PAN: {client.client_name} (ID: {client.id}, PAN: {client.pan_no})")
                    except Client.DoesNotExist:
                        # Fallback: Try PAN only (in case client name has minor differences)
                        try:
                            client = Client.objects.get(pan_no__iexact=pan_no)
                            print(f"Found client by PAN only: {client.client_name} (ID: {client.id}, PAN: {client.pan_no})")
                            print(f"WARNING: Client name mismatch - Sheet: '{client_name}', DB: '{client.client_name}'")
                        except Client.DoesNotExist:
                            pass
                        except Client.MultipleObjectsReturned:
                            skipped_records.append({
                                'client_name': client_name,
                                'pan_no': pan_no,
                                'gst_number': gst_number,
                                'reason': 'Multiple clients found with same PAN'
                            })
                            continue
                    except Client.MultipleObjectsReturned:
                        skipped_records.append({
                            'client_name': client_name,
                            'pan_no': pan_no,
                            'gst_number': gst_number,
                            'reason': 'Multiple clients found with same name and PAN'
                        })
                        continue
                else:
                    # Fallback: Client Name only (less accurate, following existing admin logic)
                    clients = Client.objects.filter(client_name__iexact=client_name)
                    if not clients.exists():
                        pass  # Will be handled below
                    elif clients.count() > 1:
                        skipped_records.append({
                            'client_name': client_name,
                            'pan_no': pan_no,
                            'gst_number': gst_number,
                            'reason': 'Multiple clients found with same name (PAN required for unique identification)'
                        })
                        continue
                    else:
                        client = clients.first()
                        print(f"Found client by name only: {client.client_name} (ID: {client.id}, PAN: {client.pan_no})")
                
                if not client:
                    reason = 'Client not found'
                    if pan_no:
                        reason += ' (name + PAN mismatch)'
                    skipped_records.append({
                        'client_name': client_name,
                        'pan_no': pan_no,
                        'gst_number': gst_number,
                        'reason': reason
                    })
                    continue
                
                # Check if client already has THIS SPECIFIC GST number
                # One client can have multiple GST numbers, so we check for exact match
                existing_gst = GSTDetails.objects.filter(
                    client=client,
                    gst_number=gst_number
                ).first()
                
                if existing_gst:
                    # Client already has this exact GST number - SKIP (duplicate)
                    print(f"Client {client.client_name} already has GST number {gst_number}, skipping duplicate")
                    skipped_records.append({
                        'client_name': client_name,
                        'pan_no': pan_no or client.pan_no,
                        'gst_number': gst_number,
                        'reason': 'Duplicate GST number (already exists for this client)'
                    })
                    continue
                else:
                    # Client doesn't have this GST number - ADD it
                    # (Client may have other GST numbers, but not this one)
                    existing_gst_count = GSTDetails.objects.filter(client=client).count()
                    print(f"Client {client.client_name} has {existing_gst_count} existing GST(s), adding new GST: {gst_number}")
                    
                    try:
                        # Create new GST details
                        new_gst = GSTDetails.objects.create(
                            client=client,
                            gst_number=gst_number,
                            registered_address=registered_address or client.address_line1,
                            state=state_code,
                            gst_scheme_type=gst_scheme_type,
                            status=status,
                            created_by=request.user
                        )
                        print(f"Created new GST ID: {new_gst.id} for {client.client_name}")
                        
                        added_records.append({
                            'client_name': client_name,
                            'pan_no': pan_no or client.pan_no,
                            'gst_number': gst_number,
                            'state': state_code,
                            'state_name': state_name,
                            'scheme_type': gst_scheme_type,
                            'status': status
                        })
                    except Exception as create_error:
                        print(f"Error creating GST for {client.client_name}: {str(create_error)}")
                        import traceback
                        traceback.print_exc()
                        skipped_records.append({
                            'client_name': client_name,
                            'pan_no': pan_no,
                            'gst_number': gst_number,
                            'reason': f'Create failed: {str(create_error)}'
                        })
                        continue
        
        print(f"Processing complete. Added: {len(added_records)}, Skipped: {len(skipped_records)}")
        
        result = {
            'success': True,
            'total_rows': total_rows,
            'added': len(added_records),
            'skipped': len(skipped_records),
            'added_records': added_records,
            'skipped_records': skipped_records
        }
        
        print(f"Returning result: {result}")
        return JsonResponse(result)
    
    except Exception as e:
        print(f"Error in bulk_import_gst: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': f'Error processing file: {str(e)}'
        })


@login_required
@user_passes_test(is_superuser_or_admin)
def download_gst_template(request):
    """Download Excel template for GST import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "GST Import Template"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['Client Name', 'PAN No', 'GST Number', 'Registered Address', 'State', 'GST Scheme Type', 'Status']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Sample data with state names (not codes)
    sample_data = [
        ['ABC Traders', 'ABCDE1234F', '27AABCU9603R1ZM', '123 Main St, Mumbai', 'Maharashtra', 'Regular', 'Active'],
        ['XYZ Industries', 'XYZAB5678G', '19AADCS1234F1Z5', '456 Park Ave, Srinagar', 'Jammu and Kashmir', 'Composition', 'Active'],
        ['PQR Enterprises', 'PQRST9876H', '09PQRST9876H1Z2', '', '', 'Regular', 'Active'],
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    
    instructions_data = [
        ['Field Name', 'Required', 'Description', 'Example'],
        ['Client Name', 'Yes', 'Client name (must match existing client)', 'ABC Traders'],
        ['PAN No', 'Recommended', 'PAN number for unique client identification', 'ABCDE1234F'],
        ['GST Number', 'Yes', '15-character GST number', '27AABCU9603R1ZM'],
        ['Registered Address', 'No', 'GST registered address (optional)', '123 Main Street, Mumbai'],
        ['State', 'No', 'State name or code (auto-calculated from GST if blank)', 'Maharashtra or 27'],
        ['GST Scheme Type', 'No', 'Regular / Composition / QRMP (default: Regular)', 'Regular'],
        ['Status', 'No', 'Active / Closed (default: Active)', 'Active'],
        ['', '', '', ''],
        ['STATE VALIDATION LOGIC:', '', '', ''],
        ['1. State code is extracted from GST number (first 2 digits)', '', '', ''],
        ['2. If State column is BLANK → uses GST-derived state code', '', '', ''],
        ['3. If State column has STATE NAME → converts to code and validates', '', '', ''],
        ['4. If State column has STATE CODE → validates against GST number', '', '', ''],
        ['5. Mismatch between provided state and GST state = SKIPPED', '', '', ''],
        ['6. Example: GST "27AABCU9603R1ZM" → State Code 27 (Maharashtra)', '', '', ''],
        ['', '', '', ''],
        ['ENHANCED PAN-BASED VALIDATION:', '', '', ''],
        ['1. Client matching uses BOTH Client Name AND PAN No when available', '', '', ''],
        ['2. PAN No ensures unique identification even with duplicate client names', '', '', ''],
        ['3. If PAN No is provided, it must match exactly with database', '', '', ''],
        ['4. GST Number must be 15 characters in valid format', '', '', ''],
        ['', '', '', ''],
        ['MULTIPLE GST SUPPORT:', '', '', ''],
        ['1. One client can have multiple GST registrations', '', '', ''],
        ['2. If client already has the SAME GST number → SKIPPED (duplicate)', '', '', ''],
        ['3. If client has DIFFERENT GST numbers → NEW GST is ADDED', '', '', ''],
        ['4. GST numbers are never updated, only added', '', '', ''],
        ['', '', '', ''],
        ['COMMON GST STATE CODES:', '', '', ''],
        ['01=West Bengal, 07=Delhi, 09=Uttar Pradesh, 19=J&K', '', '', ''],
        ['24=Gujarat, 27=Maharashtra, 29=Karnataka, 33=Tamil Nadu', '', '', ''],
        ['36=Telangana, 37=Andhra Pradesh, 32=Kerala, 30=Goa', '', '', ''],
    ]
    
    for row_num, row_data in enumerate(instructions_data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_instructions.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:  # Header row
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
    
    # Adjust column widths
    for ws_sheet in [ws, ws_instructions]:
        for col in ws_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_sheet.column_dimensions[column].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=GST_Import_Template.xlsx'
    
    wb.save(response)
    return response
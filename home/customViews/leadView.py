from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from datetime import timedelta
from django.contrib.auth.models import User
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from home.models import Lead, LeadCallLog
from home.forms import LeadForm


def get_accessible_leads(user):
    """
    Returns leads accessible to the user based on their role:
    - Admin/Superuser: All leads
    - Partner: All leads (view-only)
    - Branch Manager: Leads from their branch + assigned to them + created by them
    - Staff: Leads assigned to them OR created by them
    """
    if user.is_superuser:
        return Lead.objects.all()

    try:
        employee = user.employee
        role = employee.role

        if role == 'ADMIN':
            # Admin can see all leads
            return Lead.objects.all()
        elif role == 'PARTNER':
            # Partner can see all leads (view-only access)
            return Lead.objects.all()
        elif role == 'BRANCH_MANAGER':
            # Branch manager can see leads from their branch + assigned to them + created by them
            office = employee.office_location
            if office:
                # Get leads created by users in the same office OR assigned to this manager OR created by this manager
                return Lead.objects.filter(
                    Q(created_by__employee__office_location=office) |
                    Q(assigned_to=user) |
                    Q(created_by=user)
                ).distinct()
            else:
                # If no office, show assigned + created leads
                return Lead.objects.filter(
                    Q(assigned_to=user) |
                    Q(created_by=user)
                ).distinct()
        else:
            # Staff can see leads assigned to them OR created by them
            return Lead.objects.filter(
                Q(assigned_to=user) |
                Q(created_by=user)
            ).distinct()
    except:
        # If no employee profile, only show assigned + created leads
        return Lead.objects.filter(
            Q(assigned_to=user) |
            Q(created_by=user)
        ).distinct()


@login_required
def lead_list_view(request):
    """List all leads with filtering based on user role"""
    leads = get_accessible_leads(request.user).select_related('created_by').prefetch_related('assigned_to')

    # Apply filters
    search_query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()
    assigned_filter = request.GET.get('assigned_to', '').strip()

    if search_query:
        leads = leads.filter(
            Q(lead_name__icontains=search_query) |
            Q(full_name__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    if status_filter:
        leads = leads.filter(status=status_filter)

    if assigned_filter:
        leads = leads.filter(assigned_to__id=assigned_filter)

    # Get filter choices
    from django.contrib.auth.models import User
    assigned_users = User.objects.filter(is_active=True).order_by('username')

    # Check if user is a partner (view-only access)
    is_partner = False
    if hasattr(request.user, 'employee'):
        is_partner = request.user.employee.role == 'PARTNER'

    context = {
        'leads': leads,
        'status_choices': Lead.STATUS_CHOICES,
        'assigned_users': assigned_users,
        'search_query': search_query,
        'status_filter': status_filter,
        'assigned_filter': assigned_filter,
        'is_partner': is_partner,
    }

    return render(request, 'leads/lead_list.html', context)


@login_required
def add_lead_view(request):
    """Add a new lead and auto-assign to creator"""
    if request.method == 'POST':
        form = LeadForm(request.POST)
        if form.is_valid():
            lead = form.save(commit=False)
            lead.created_by = request.user
            lead.save()
            # Auto-assign the lead to the creator
            lead.assigned_to.add(request.user)
            messages.success(request, f'Lead "{lead.lead_name}" created successfully!')
            return redirect('lead_list')
    else:
        form = LeadForm()

    context = {
        'form': form,
        'page_title': 'Add New Lead',
    }
    return render(request, 'leads/add_lead.html', context)


@login_required
def edit_lead_view(request, lead_id):
    """Edit an existing lead (only if status is New, Contacted, or Qualified)"""
    lead = get_object_or_404(Lead, id=lead_id)
    
    # Check if user is a partner - apply restricted edit permissions
    if hasattr(request.user, 'employee') and request.user.employee.role == 'PARTNER':
        # Partner can edit ONLY IF they onboarded the lead OR are assigned to the lead
        if lead.created_by != request.user and request.user not in lead.assigned_to.all():
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden("You are not allowed to perform this action")

    # Check if user has access to this lead
    accessible_leads = get_accessible_leads(request.user)
    if not accessible_leads.filter(id=lead.id).exists():
        messages.error(request, 'You do not have permission to edit this lead.')
        return redirect('lead_list')

    # Check if lead is editable
    if not lead.is_editable():
        messages.error(request, f'Cannot edit lead in "{lead.status}" status.')
        return redirect('lead_detail', lead_id=lead.id)

    if request.method == 'POST':
        form = LeadForm(request.POST, instance=lead)
        if form.is_valid():
            form.save()
            messages.success(request, f'Lead "{lead.lead_name}" updated successfully!')
            return redirect('lead_detail', lead_id=lead.id)
    else:
        form = LeadForm(instance=lead)

    context = {
        'form': form,
        'lead': lead,
        'page_title': 'Edit Lead',
    }
    return render(request, 'leads/edit_lead.html', context)


@login_required
def lead_detail_view(request, lead_id):
    """View lead details"""
    lead = get_object_or_404(Lead.objects.select_related('created_by', 'client').prefetch_related('assigned_to'),
                             id=lead_id)

    # Check if user has access to this lead
    accessible_leads = get_accessible_leads(request.user)
    if not accessible_leads.filter(id=lead.id).exists():
        messages.error(request, 'You do not have permission to view this lead.')
        return redirect('lead_list')

    # Get call logs for this lead
    call_logs = lead.call_logs.select_related('called_by_user', 'called_to_user', 'created_by').all()

    # Check if user is a partner (view-only access)
    is_partner = False
    can_edit_lead = True
    if hasattr(request.user, 'employee'):
        is_partner = request.user.employee.role == 'PARTNER'
        if is_partner:
            # Partner can edit ONLY IF they onboarded the lead OR are assigned to the lead
            can_edit_lead = (lead.created_by == request.user or request.user in lead.assigned_to.all())

    # Check if user can add call logs (creator, assigned users, branch manager, admin, and Partners with edit access)
    can_add_call_log = False
    if request.user.is_superuser:
        can_add_call_log = True
    elif is_partner:
        # Partners CAN add call logs if they have edit access to the lead
        can_add_call_log = can_edit_lead
    elif lead.created_by == request.user or request.user in lead.assigned_to.all():
        can_add_call_log = True
    else:
        try:
            employee = request.user.employee
            if employee.role in ['ADMIN', 'BRANCH_MANAGER']:
                # Branch manager can add if lead is from their office
                if employee.role == 'BRANCH_MANAGER' and employee.office_location:
                    if lead.created_by and hasattr(lead.created_by, 'employee'):
                        if lead.created_by.employee.office_location == employee.office_location:
                            can_add_call_log = True
                else:
                    can_add_call_log = True
        except:
            pass

    # Get authorized users for dropdown (creator + assigned users + branch manager if applicable)
    authorized_users = set()
    if lead.created_by:
        authorized_users.add(lead.created_by)
    authorized_users.update(lead.assigned_to.all())

    # Add branch manager if applicable
    if lead.created_by and hasattr(lead.created_by, 'employee'):
        office = lead.created_by.employee.office_location
        if office:
            from django.contrib.auth.models import User
            branch_managers = User.objects.filter(
                employee__role='BRANCH_MANAGER',
                employee__office_location=office
            )
            authorized_users.update(branch_managers)

    context = {
        'lead': lead,
        'call_logs': call_logs,
        'can_add_call_log': can_add_call_log,
        'authorized_users': sorted(authorized_users, key=lambda u: u.get_full_name() or u.username),
        'today': timezone.now().date(),
        'is_partner': is_partner,
        'can_edit_lead': can_edit_lead,
    }
    return render(request, 'leads/lead_detail.html', context)


@login_required
@require_POST
def mark_lead_contacted(request, lead_id):
    """Mark lead as Contacted"""
    # Check if user is a partner - deny access
    if hasattr(request.user, 'employee') and request.user.employee.role == 'PARTNER':
        messages.error(request, 'You do not have permission to change lead status.')
        return redirect('lead_detail', lead_id=lead_id)

    lead = get_object_or_404(Lead, id=lead_id)

    # Check if user has access to this lead
    accessible_leads = get_accessible_leads(request.user)
    if not accessible_leads.filter(id=lead.id).exists():
        messages.error(request, 'You do not have permission to update this lead.')
        return redirect('lead_list')

    if lead.status != 'New':
        messages.error(request, 'Can only mark New leads as Contacted.')
        return redirect('lead_detail', lead_id=lead.id)

    lead.status = 'Contacted'
    lead.save()
    messages.success(request, f'Lead "{lead.lead_name}" marked as Contacted.')
    return redirect('lead_detail', lead_id=lead.id)


@login_required
@require_POST
def mark_lead_qualified(request, lead_id):
    """Mark lead as Qualified"""
    # Check if user is a partner - deny access
    if hasattr(request.user, 'employee') and request.user.employee.role == 'PARTNER':
        messages.error(request, 'You do not have permission to change lead status.')
        return redirect('lead_detail', lead_id=lead_id)

    lead = get_object_or_404(Lead, id=lead_id)

    # Check if user has access to this lead
    accessible_leads = get_accessible_leads(request.user)
    if not accessible_leads.filter(id=lead.id).exists():
        messages.error(request, 'You do not have permission to update this lead.')
        return redirect('lead_list')

    if lead.status != 'Contacted':
        messages.error(request, 'Can only mark Contacted leads as Qualified.')
        return redirect('lead_detail', lead_id=lead.id)

    lead.status = 'Qualified'
    lead.save()
    messages.success(request, f'Lead "{lead.lead_name}" marked as Qualified.')
    return redirect('lead_detail', lead_id=lead.id)


@login_required
@require_POST
def mark_lead_lost(request, lead_id):
    """Mark lead as Lost with reason"""
    # Check if user is a partner - deny access
    if hasattr(request.user, 'employee') and request.user.employee.role == 'PARTNER':
        messages.error(request, 'You do not have permission to change lead status.')
        return redirect('lead_detail', lead_id=lead_id)

    lead = get_object_or_404(Lead, id=lead_id)

    # Check if user has access to this lead
    accessible_leads = get_accessible_leads(request.user)
    if not accessible_leads.filter(id=lead.id).exists():
        messages.error(request, 'You do not have permission to update this lead.')
        return redirect('lead_list')

    # Can mark as Lost from New, Contacted, or Qualified
    if lead.status not in ['New', 'Contacted', 'Qualified']:
        messages.error(request, f'Cannot mark lead as Lost from "{lead.status}" status.')
        return redirect('lead_detail', lead_id=lead.id)

    reason = request.POST.get('reason', '').strip()
    if not reason:
        messages.error(request, 'Please provide a reason for marking the lead as Lost.')
        return redirect('lead_detail', lead_id=lead.id)

    lead.status = 'Lost'
    lead.remarks = reason
    lead.actual_closure_date = timezone.now().date()
    lead.save()

    messages.success(request, f'Lead "{lead.lead_name}" marked as Lost.')
    return redirect('lead_detail', lead_id=lead.id)


@login_required
def convert_lead_view(request, lead_id):
    """Convert lead to client - redirect to onboarding with prefilled data"""
    # Check if user is a partner - deny access
    if hasattr(request.user, 'employee') and request.user.employee.role == 'PARTNER':
        messages.error(request, 'You do not have permission to convert leads.')
        return redirect('lead_detail', lead_id=lead_id)

    lead = get_object_or_404(Lead, id=lead_id)

    # Check if user has access to this lead
    accessible_leads = get_accessible_leads(request.user)
    if not accessible_leads.filter(id=lead.id).exists():
        messages.error(request, 'You do not have permission to convert this lead.')
        return redirect('lead_list')

    # Check if lead is qualified
    if lead.status != 'Qualified':
        messages.error(request, 'Can only convert Qualified leads.')
        return redirect('lead_detail', lead_id=lead.id)

    # Check if already converted
    if lead.client:
        messages.warning(request, f'This lead has already been converted to client: {lead.client.client_name}')
        return redirect('client_details', client_id=lead.client.id)

    # Redirect to onboarding form with lead_id
    messages.info(request, f'Converting lead "{lead.lead_name}" to client. Please complete the onboarding form.')
    return redirect(f'/onboard/?lead_id={lead.id}')


@login_required
@require_POST
def add_lead_call_log(request, lead_id):
    """Add a new call log for a lead - Employee always calls Lead"""
    lead = get_object_or_404(Lead, id=lead_id)

    # Check if user has access to this lead
    accessible_leads = get_accessible_leads(request.user)
    if not accessible_leads.filter(id=lead.id).exists():
        return JsonResponse({'success': False, 'error': 'You do not have permission to add call logs for this lead.'},
                            status=403)

    # Verify user can add call logs
    can_add = False
    if request.user.is_superuser:
        can_add = True
    elif lead.created_by == request.user or request.user in lead.assigned_to.all():
        can_add = True
    else:
        try:
            employee = request.user.employee
            # Partners can add call logs if they have edit access
            if employee.role == 'PARTNER':
                can_add = (lead.created_by == request.user or request.user in lead.assigned_to.all())
            elif employee.role in ['ADMIN', 'BRANCH_MANAGER']:
                if employee.role == 'BRANCH_MANAGER' and employee.office_location:
                    if lead.created_by and hasattr(lead.created_by, 'employee'):
                        if lead.created_by.employee.office_location == employee.office_location:
                            can_add = True
                else:
                    can_add = True
        except:
            pass

    if not can_add:
        return JsonResponse({'success': False, 'error': 'You do not have permission to add call logs.'}, status=403)

    # Get form data
    call_date = request.POST.get('call_date')
    description = request.POST.get('description', '').strip()

    # Validate inputs
    if not call_date:
        return JsonResponse({'success': False, 'error': 'Please select call date.'}, status=400)

    if not description:
        return JsonResponse({'success': False, 'error': 'Please provide a description of the call.'}, status=400)

    # Validate date format and ensure it's not in the future
    try:
        from datetime import datetime
        call_date_obj = datetime.strptime(call_date, '%Y-%m-%d').date()
        today = datetime.now().date()
        if call_date_obj > today:
            return JsonResponse({'success': False, 'error': 'Call date cannot be in the future.'}, status=400)
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Invalid date format.'}, status=400)

    # Create call log - Employee (logged-in user) always calls Lead
    try:
        call_log = LeadCallLog.objects.create(
            lead=lead,
            called_by_user=request.user,  # Employee who made the call
            called_to_user=None,  # NULL means Lead received the call
            call_date=call_date_obj,
            description=description,
            created_by=request.user
        )

        messages.success(request, 'Call log added successfully!')
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error creating call log: {str(e)}'}, status=500)



@login_required
def export_leads_to_excel(request):
    """Export filtered leads to Excel with all details"""
    # Get filtered leads using the same logic as lead_list_view
    leads = get_accessible_leads(request.user).select_related('created_by', 'client').prefetch_related('assigned_to')

    # Apply the same filters as in lead_list_view
    search_query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()
    assigned_filter = request.GET.get('assigned_to', '').strip()

    if search_query:
        leads = leads.filter(
            Q(lead_name__icontains=search_query) |
            Q(full_name__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    if status_filter:
        leads = leads.filter(status=status_filter)

    if assigned_filter:
        leads = leads.filter(assigned_to__id=assigned_filter)

    # Check if there are any leads to export
    if not leads.exists():
        messages.error(request, 'No leads to export.')
        return redirect('lead_list')

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Leads Export'

    # Define headers
    headers = [
        '#', 'Lead Name', 'Full Name', 'Email', 'Phone Number', 
        'Requirements', 'Lead Value', 'Expected Closure Date', 
        'Actual Closure Date', 'Status', 'Remarks', 
        'Assigned To', 'Created By', 'Created At', 'Updated At',
        'Converted Client'
    ]

    # Style for header row
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Set column widths
    column_widths = {
        'A': 8,   # #
        'B': 25,  # Lead Name
        'C': 25,  # Full Name
        'D': 30,  # Email
        'E': 18,  # Phone Number
        'F': 40,  # Requirements
        'G': 15,  # Lead Value
        'H': 20,  # Expected Closure Date
        'I': 20,  # Actual Closure Date
        'J': 15,  # Status
        'K': 40,  # Remarks
        'L': 30,  # Assigned To
        'M': 20,  # Created By
        'N': 20,  # Created At
        'O': 20,  # Updated At
        'P': 25,  # Converted Client
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Write data rows
    for idx, lead in enumerate(leads, start=1):
        row_num = idx + 1
        
        # Get assigned users as comma-separated string
        assigned_users = ', '.join([
            user.get_full_name() or user.username 
            for user in lead.assigned_to.all()
        ]) if lead.assigned_to.exists() else 'Unassigned'

        # Get created by user
        created_by = lead.created_by.get_full_name() or lead.created_by.username if lead.created_by else 'N/A'

        # Get converted client name
        converted_client = lead.client.client_name if lead.client else 'N/A'

        # Format dates
        expected_closure = lead.expected_closure_date.strftime('%d-%m-%Y') if lead.expected_closure_date else 'N/A'
        actual_closure = lead.actual_closure_date.strftime('%d-%m-%Y') if lead.actual_closure_date else 'N/A'
        created_at = lead.created_at.strftime('%d-%m-%Y %H:%M') if lead.created_at else 'N/A'
        updated_at = lead.updated_at.strftime('%d-%m-%Y %H:%M') if lead.updated_at else 'N/A'

        # Format lead value
        lead_value = f'₹{lead.lead_value:,.2f}' if lead.lead_value else 'N/A'

        # Write row data
        row_data = [
            idx,
            lead.lead_name or 'N/A',
            lead.full_name or 'N/A',
            lead.email or 'N/A',
            lead.phone_number or 'N/A',
            lead.requirements or 'N/A',
            lead_value,
            expected_closure,
            actual_closure,
            lead.status or 'N/A',
            lead.remarks or 'N/A',
            assigned_users,
            created_by,
            created_at,
            updated_at,
            converted_client,
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename with timestamp
    from datetime import datetime
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'Leads_Export_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save workbook to response
    wb.save(response)

    return response

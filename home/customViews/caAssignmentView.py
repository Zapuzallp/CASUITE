"""
Bulk Client CA Assignment View
Allows superusers/admins to assign or update Chartered Accountants for multiple clients
"""
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.contrib import messages
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from io import BytesIO

from home.models import Client, User, Employee


def is_superuser_or_admin(user):
    """Check if user is superuser or admin"""
    if user.is_superuser:
        return True
    try:
        return hasattr(user, 'employee') and user.employee.role == 'ADMIN'
    except:
        return False


@login_required
@user_passes_test(is_superuser_or_admin, login_url='/accounts/login/')
def assign_ca_view(request):
    """Main view for CA assignment page"""
    context = {
        'page_title': 'Assign / Update CA',
    }
    return render(request, 'ca_assignment/assign_ca.html', context)


@login_required
@user_passes_test(is_superuser_or_admin, login_url='/accounts/login/')
def search_clients_ajax(request):
    """AJAX endpoint to search clients by name or PAN"""
    q = request.GET.get('q', '').strip()
    
    print(f"Client search query: '{q}'")  # Debug
    
    if len(q) < 2:
        print("Query too short, returning empty")
        return JsonResponse([], safe=False)
    
    clients = Client.objects.filter(
        Q(client_name__icontains=q) | Q(pan_no__icontains=q)
    )[:20]
    
    data = []
    for c in clients:
        data.append({
            "id": c.id,
            "text": f"{c.client_name} ({c.pan_no or 'No PAN'})"
        })
    
    print(f"Found {len(data)} clients")  # Debug
    return JsonResponse(data, safe=False)


@login_required
@user_passes_test(is_superuser_or_admin, login_url='/accounts/login/')
def search_employees_ajax(request):
    """AJAX endpoint to search employees (potential CAs)"""
    q = request.GET.get('q', '').strip()
    
    print(f"Employee search query: '{q}'")  # Debug
    
    if len(q) < 2:
        print("Query too short, returning empty")
        return JsonResponse([], safe=False)
    
    employees = Employee.objects.filter(
        Q(user__first_name__icontains=q) |
        Q(user__last_name__icontains=q) |
        Q(user__username__icontains=q)
    ).select_related('user')[:20]
    
    data = []
    for emp in employees:
        full_name = emp.user.get_full_name() or emp.user.username
        data.append({
            "id": emp.user.id,
            "text": f"{full_name} ({emp.user.username})"
        })
    
    print(f"Found {len(data)} employees")  # Debug
    return JsonResponse(data, safe=False)


@login_required
@user_passes_test(is_superuser_or_admin)
def manual_assign_ca(request):
    """Handle manual CA assignment for multiple clients"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    client_ids = request.POST.getlist('client_ids[]')
    ca_id = request.POST.get('ca_id')
    
    if not client_ids or not ca_id:
        return JsonResponse({
            'success': False,
            'message': 'Please select both clients and CA'
        })
    
    try:
        ca_user = User.objects.get(id=ca_id)
    except User.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Selected CA not found'
        })
    
    assigned_count = 0
    updated_count = 0
    
    with transaction.atomic():
        for client_id in client_ids:
            try:
                client = Client.objects.get(id=client_id)
                if client.assigned_ca:
                    updated_count += 1
                else:
                    assigned_count += 1
                client.assigned_ca = ca_user
                client.save()
            except Client.DoesNotExist:
                continue
    
    return JsonResponse({
        'success': True,
        'assigned': assigned_count,
        'updated': updated_count,
        'total': assigned_count + updated_count
    })


@login_required
@user_passes_test(is_superuser_or_admin)
def bulk_import_ca(request):
    """Handle bulk CA assignment via Excel/CSV upload"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No file uploaded'})
    
    file = request.FILES['file']
    
    # Validate file extension
    if not (file.name.endswith('.xlsx') or file.name.endswith('.xls') or file.name.endswith('.csv')):
        return JsonResponse({
            'success': False,
            'message': 'Invalid file format. Please upload Excel or CSV file'
        })
    
    try:
        # Read file
        if file.name.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        # Validate columns
        required_columns = ['Client Name', 'PAN No', 'Employee Name']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return JsonResponse({
                'success': False,
                'message': f'Missing required columns: {", ".join(missing_columns)}'
            })
        
        # Process rows
        total_rows = len(df)
        assigned_count = 0
        updated_count = 0
        skipped_rows = []
        
        with transaction.atomic():
            for index, row in df.iterrows():
                pan_no = str(row['PAN No']).strip() if pd.notna(row['PAN No']) else ''
                employee_name = str(row['Employee Name']).strip() if pd.notna(row['Employee Name']) else ''
                client_name = str(row['Client Name']).strip() if pd.notna(row['Client Name']) else ''
                
                # Skip if employee name is missing
                if not employee_name:
                    skipped_rows.append({
                        'client_name': client_name,
                        'pan_no': pan_no,
                        'reason': 'Employee missing'
                    })
                    continue
                
                # Find client by PAN
                try:
                    client = Client.objects.get(pan_no__iexact=pan_no)
                except Client.DoesNotExist:
                    skipped_rows.append({
                        'client_name': client_name,
                        'pan_no': pan_no,
                        'reason': 'Client not found'
                    })
                    continue
                
                # Find employee by name
                employee_parts = employee_name.split()
                if len(employee_parts) >= 2:
                    first_name = employee_parts[0]
                    last_name = ' '.join(employee_parts[1:])
                    employee_query = Q(first_name__iexact=first_name, last_name__iexact=last_name)
                else:
                    employee_query = Q(username__iexact=employee_name) | Q(first_name__iexact=employee_name)
                
                try:
                    ca_user = User.objects.filter(employee__isnull=False).filter(employee_query).first()
                    if not ca_user:
                        raise User.DoesNotExist
                except User.DoesNotExist:
                    skipped_rows.append({
                        'client_name': client_name,
                        'pan_no': pan_no,
                        'reason': 'Employee not found'
                    })
                    continue
                
                # Assign or update CA
                if client.assigned_ca:
                    updated_count += 1
                else:
                    assigned_count += 1
                
                client.assigned_ca = ca_user
                client.save()
        
        return JsonResponse({
            'success': True,
            'total_rows': total_rows,
            'assigned': assigned_count,
            'updated': updated_count,
            'skipped': len(skipped_rows),
            'skipped_rows': skipped_rows
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error processing file: {str(e)}'
        })


@login_required
@user_passes_test(is_superuser_or_admin)
def download_demo_template(request):
    """Download demo Excel template for CA assignment"""
    wb = Workbook()
    ws = wb.active
    ws.title = "CA Assignment Template"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['Client Name', 'PAN No', 'Employee Name']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Sample data
    sample_data = [
        ['ABC Pvt Ltd', 'ABCDE1234F', 'John Doe'],
        ['XYZ Company', 'XYZAB5678G', 'Jane Smith'],
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=CA_Assignment_Template.xlsx'
    
    wb.save(response)
    return response

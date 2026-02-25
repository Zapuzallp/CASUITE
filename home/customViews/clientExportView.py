"""
Dynamic Multi-Model Export to Excel - Filter-Agnostic, Pagination-Safe, RBAC-Compliant

This module provides a configuration-driven export system that can handle multiple models.
To add a new model for export, simply add a new configuration to EXPORT_CONFIGS.
"""
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from home.models import Client, Task
from home.clients.client_access import get_accessible_clients


# ========================================================================================
# HELPER FUNCTIONS
# ========================================================================================

def is_gst_number(query):
    """Check if the query string matches GST number format."""
    import re
    gst_pattern = r'^\d{2}[A-Z]{5}\d{4}[A-Z]{1}[A-Z\d]{1}Z[A-Z\d]{1}$'
    return bool(re.match(gst_pattern, query.upper()))


# ========================================================================================
# QUERYSET FUNCTIONS (One per model)
# ========================================================================================

def get_filtered_clients(request):
    """
    Centralized filtering logic for clients.
    Returns filtered queryset BEFORE pagination.
    """
    user = request.user

    # 1. Start with role-based access control
    qs = get_accessible_clients(user)

    # 2. Extract GET parameters
    search_query = request.GET.get('q')
    filter_status = request.GET.get('status')
    filter_type = request.GET.get('client_type')
    filter_structure = request.GET.get('business_structure')
    filter_office = request.GET.get('office')
    filter_service = request.GET.get('service_type')
    filter_assigned_to = request.GET.get('assigned_to')
    filter_custom_view = request.GET.get('custom_view')

    # 3. Apply filters dynamically
    if search_query:
        if is_gst_number(search_query):
            qs = qs.filter(gst_details__gst_number__icontains=search_query).distinct()
        else:
            qs = qs.filter(
                Q(client_name__icontains=search_query) |
                Q(pan_no__icontains=search_query) |
                Q(file_number__icontains=search_query) |
                Q(primary_contact_name__icontains=search_query)
            )

    if filter_status:
        qs = qs.filter(status=filter_status)

    if filter_type:
        qs = qs.filter(client_type=filter_type)

    if filter_structure:
        qs = qs.filter(business_structure=filter_structure)

    if filter_office:
        qs = qs.filter(office_location_id=filter_office)

    if filter_service:
        qs = qs.filter(tasks__service_type=filter_service).distinct()

    if filter_assigned_to:
        qs = qs.filter(assigned_ca_id=filter_assigned_to)

    # Apply Custom View Filter
    if filter_custom_view:
        if filter_custom_view == 'aadhaar_mobile_linked':
            qs = qs.filter(
                aadhar__isnull=False,
                phone_number__isnull=False
            ).exclude(aadhar='').exclude(phone_number='')
        elif filter_custom_view == 'gst_enabled':
            qs = qs.filter(gst_details__isnull=False).distinct()
        elif filter_custom_view == 'director_din_valid':
            qs = qs.filter(din_no__isnull=False).exclude(din_no='')

    return qs.order_by('-created_at')


# ========================================================================================
# FIELD CONFIGURATIONS (One per model)
# ========================================================================================

CLIENT_FIELDS = {
    # Client Model Fields
    'file_number': {
        'label': 'File Number',
        'group': 'client',
        'accessor': lambda c: c.file_number or ''
    },
    'client_name': {
        'label': 'Client Name',
        'group': 'client',
        'accessor': lambda c: c.client_name
    },
    'primary_contact_name': {
        'label': 'Primary Contact Name',
        'group': 'client',
        'accessor': lambda c: c.primary_contact_name
    },
    'pan_no': {
        'label': 'PAN Number',
        'group': 'client',
        'accessor': lambda c: c.pan_no
    },
    'aadhar': {
        'label': 'Aadhar Number',
        'group': 'client',
        'accessor': lambda c: c.aadhar or ''
    },
    'aadhar_linked_mobile': {
        'label': 'Aadhar Linked Mobile',
        'group': 'client',
        'accessor': lambda c: 'Yes' if c.aadhar_linked_mobile else 'No'
    },
    'din_no': {
        'label': 'DIN Number',
        'group': 'client',
        'accessor': lambda c: c.din_no or ''
    },
    'tan_no': {
        'label': 'TAN Number',
        'group': 'client',
        'accessor': lambda c: c.tan_no or ''
    },
    'email': {
        'label': 'Email',
        'group': 'client',
        'accessor': lambda c: c.email
    },
    'phone_number': {
        'label': 'Phone Number',
        'group': 'client',
        'accessor': lambda c: c.phone_number
    },
    'father_name': {
        'label': 'Father Name',
        'group': 'client',
        'accessor': lambda c: c.father_name or ''
    },
    'address_line1': {
        'label': 'Address',
        'group': 'client',
        'accessor': lambda c: c.address_line1
    },
    'postal_code': {
        'label': 'Postal Code',
        'group': 'client',
        'accessor': lambda c: c.postal_code
    },
    'city': {
        'label': 'City',
        'group': 'client',
        'accessor': lambda c: c.city
    },
    'state': {
        'label': 'State',
        'group': 'client',
        'accessor': lambda c: c.get_state_display() if c.state else ''
    },
    'country': {
        'label': 'Country',
        'group': 'client',
        'accessor': lambda c: c.country
    },
    'office_location': {
        'label': 'Office Location',
        'group': 'client',
        'accessor': lambda c: c.office_location.office_name if c.office_location else ''
    },
    'date_of_engagement': {
        'label': 'Date of Engagement',
        'group': 'client',
        'accessor': lambda c: c.date_of_engagement.strftime('%Y-%m-%d') if c.date_of_engagement else ''
    },
    'assigned_ca': {
        'label': 'Assigned CA',
        'group': 'client',
        'accessor': lambda c: c.assigned_ca.get_full_name() or c.assigned_ca.username if c.assigned_ca else ''
    },
    'client_type': {
        'label': 'Client Type',
        'group': 'client',
        'accessor': lambda c: c.get_client_type_display() if c.client_type else ''
    },
    'business_structure': {
        'label': 'Business Structure',
        'group': 'client',
        'accessor': lambda c: c.get_business_structure_display() if c.business_structure else ''
    },
    'status': {
        'label': 'Status',
        'group': 'client',
        'accessor': lambda c: c.get_status_display() if c.status else ''
    },
    'remarks': {
        'label': 'Remarks',
        'group': 'client',
        'accessor': lambda c: c.remarks or ''
    },
    'created_by': {
        'label': 'Created By',
        'group': 'client',
        'accessor': lambda c: c.created_by.username if c.created_by else ''
    },
    'created_at': {
        'label': 'Created At',
        'group': 'client',
        'accessor': lambda c: c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else ''
    },
    'updated_at': {
        'label': 'Updated At',
        'group': 'client',
        'accessor': lambda c: c.updated_at.strftime('%Y-%m-%d %H:%M') if c.updated_at else ''
    },

    # ClientBusinessProfile Fields
    'registration_number': {
        'label': 'Registration Number (CIN/LLPIN)',
        'group': 'profile',
        'accessor': lambda c: c.business_profile.registration_number if hasattr(c, 'business_profile') and c.business_profile.registration_number else ''
    },
    'date_of_incorporation': {
        'label': 'Date of Incorporation',
        'group': 'profile',
        'accessor': lambda c: c.business_profile.date_of_incorporation.strftime('%Y-%m-%d') if hasattr(c, 'business_profile') and c.business_profile.date_of_incorporation else ''
    },
    'registered_office_address': {
        'label': 'Registered Office Address',
        'group': 'profile',
        'accessor': lambda c: c.business_profile.registered_office_address if hasattr(c, 'business_profile') and c.business_profile.registered_office_address else ''
    },
    'udyam_registration': {
        'label': 'Udyam Registration',
        'group': 'profile',
        'accessor': lambda c: c.business_profile.udyam_registration if hasattr(c, 'business_profile') and c.business_profile.udyam_registration else ''
    },
    'authorised_capital': {
        'label': 'Authorised Capital',
        'group': 'profile',
        'accessor': lambda c: str(c.business_profile.authorised_capital) if hasattr(c, 'business_profile') and c.business_profile.authorised_capital else ''
    },
    'paid_up_capital': {
        'label': 'Paid-up Capital',
        'group': 'profile',
        'accessor': lambda c: str(c.business_profile.paid_up_capital) if hasattr(c, 'business_profile') and c.business_profile.paid_up_capital else ''
    },
    'number_of_directors': {
        'label': 'Number of Directors',
        'group': 'profile',
        'accessor': lambda c: str(c.business_profile.number_of_directors) if hasattr(c, 'business_profile') and c.business_profile.number_of_directors else ''
    },
    'number_of_shareholders': {
        'label': 'Number of Shareholders',
        'group': 'profile',
        'accessor': lambda c: str(c.business_profile.number_of_shareholders) if hasattr(c, 'business_profile') and c.business_profile.number_of_shareholders else ''
    },
    'number_of_members': {
        'label': 'Number of Members',
        'group': 'profile',
        'accessor': lambda c: str(c.business_profile.number_of_members) if hasattr(c, 'business_profile') and c.business_profile.number_of_members else ''
    },
    'number_of_coparceners': {
        'label': 'Number of Coparceners',
        'group': 'profile',
        'accessor': lambda c: str(c.business_profile.number_of_coparceners) if hasattr(c, 'business_profile') and c.business_profile.number_of_coparceners else ''
    },
    'opc_nominee_name': {
        'label': 'OPC Nominee Name',
        'group': 'profile',
        'accessor': lambda c: c.business_profile.opc_nominee_name if hasattr(c, 'business_profile') and c.business_profile.opc_nominee_name else ''
    },
    'object_clause': {
        'label': 'Object Clause',
        'group': 'profile',
        'accessor': lambda c: c.business_profile.object_clause if hasattr(c, 'business_profile') and c.business_profile.object_clause else ''
    },
    'is_section8_license_obtained': {
        'label': 'Section 8 License Obtained',
        'group': 'profile',
        'accessor': lambda c: 'Yes' if hasattr(c, 'business_profile') and c.business_profile.is_section8_license_obtained else 'No'
    },
    'key_persons': {
        'label': 'Key Persons',
        'group': 'profile',
        'accessor': lambda c: ', '.join([kp.client_name for kp in c.business_profile.key_persons.all()]) if hasattr(c, 'business_profile') else ''
    },
    'constitution_document_1': {
        'label': 'Constitution Document 1',
        'group': 'profile',
        'accessor': lambda c: c.business_profile.constitution_document_1.name if hasattr(c, 'business_profile') and c.business_profile.constitution_document_1 else ''
    },
    'constitution_document_2': {
        'label': 'Constitution Document 2',
        'group': 'profile',
        'accessor': lambda c: c.business_profile.constitution_document_2.name if hasattr(c, 'business_profile') and c.business_profile.constitution_document_2 else ''
    },
}


# ========================================================================================
# MASTER EXPORT CONFIGURATION
# ========================================================================================
#
# HOW TO ADD A NEW MODEL FOR EXPORT:
#
# 1. Create a queryset function (like get_filtered_clients above)
# 2. Create a fields dictionary (like CLIENT_FIELDS above)
# 3. Add an entry here with:
#    - model_name: Django model class
#    - display_name: Human-readable name
#    - queryset_func: Function that returns filtered queryset
#    - fields: Dictionary of exportable fields
#    - select_related: List of relations to optimize (optional)
#    - prefetch_related: List of M2M relations to optimize (optional)
#
# Example for Task model:
# 'task': {
#     'model_name': Task,
#     'display_name': 'Tasks',
#     'queryset_func': get_filtered_tasks_for_clients,
#     'fields': TASK_FIELDS,
#     'select_related': ['client', 'created_by'],
#     'prefetch_related': ['assignees'],
# }
# ========================================================================================

EXPORT_CONFIGS = {
    'client': {
        'model_name': Client,
        'display_name': 'Clients',
        'queryset_func': get_filtered_clients,
        'fields': CLIENT_FIELDS,
        'select_related': ['assigned_ca', 'created_by', 'office_location', 'business_profile'],
        'prefetch_related': ['business_profile__key_persons'],
    },
    # Add more models here in the future
}


# ========================================================================================
# VIEWS
# ========================================================================================

@login_required
def client_export_select_columns(request):
    """
    Display column selection page for export.
    Dynamically handles any model configured in EXPORT_CONFIGS.
    """
    # Get model type from query params (default to 'client')
    model_type = request.GET.get('model', 'client')

    # Validate model type
    if model_type not in EXPORT_CONFIGS:
        from django.contrib import messages
        messages.error(request, f'Invalid export model: {model_type}')
        return redirect('clients')

    # Get configuration for this model
    config = EXPORT_CONFIGS[model_type]

    # Get filtered queryset (no pagination)
    queryset = config['queryset_func'](request)
    total_count = queryset.count()

    # Group fields by their group attribute
    grouped_fields = {}
    for field_key, field_info in config['fields'].items():
        group = field_info.get('group', 'default')
        if group not in grouped_fields:
            grouped_fields[group] = {}
        grouped_fields[group][field_key] = field_info

    context = {
        'model_type': model_type,
        'display_name': config['display_name'],
        'total_count': total_count,
        'grouped_fields': grouped_fields,
        'query_params': request.GET.urlencode(),
    }

    return render(request, 'client/export_select_columns.html', context)


@login_required
def client_export_generate(request):
    """
    Generate Excel file with selected columns.
    Dynamically handles any model configured in EXPORT_CONFIGS.
    """
    if request.method != 'POST':
        return redirect('client_export_select_columns')

    # Get model type
    model_type = request.GET.get('model', 'client')

    # Validate model type
    if model_type not in EXPORT_CONFIGS:
        from django.contrib import messages
        messages.error(request, f'Invalid export model: {model_type}')
        return redirect('clients')

    # Get configuration
    config = EXPORT_CONFIGS[model_type]

    # Get selected columns from POST
    selected_fields = request.POST.getlist('fields')

    if not selected_fields:
        from django.contrib import messages
        messages.error(request, 'Please select at least one column to export.')
        return redirect('client_export_select_columns')

    # Get filtered queryset with optimizations
    queryset = config['queryset_func'](request)

    # Apply select_related and prefetch_related if configured
    if 'select_related' in config and config['select_related']:
        queryset = queryset.select_related(*config['select_related'])

    if 'prefetch_related' in config and config['prefetch_related']:
        queryset = queryset.prefetch_related(*config['prefetch_related'])

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{config['display_name']} Export"

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write header row
    for col_idx, field_key in enumerate(selected_fields, start=1):
        if field_key in config['fields']:
            cell = ws.cell(row=1, column=col_idx)
            cell.value = config['fields'][field_key]['label']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    # Write data rows
    for row_idx, record in enumerate(queryset, start=2):
        for col_idx, field_key in enumerate(selected_fields, start=1):
            if field_key in config['fields']:
                try:
                    value = config['fields'][field_key]['accessor'](record)
                    ws.cell(row=row_idx, column=col_idx, value=value)
                except Exception:
                    # Handle any errors gracefully
                    ws.cell(row=row_idx, column=col_idx, value='')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{model_type}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response
